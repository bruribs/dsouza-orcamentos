from __future__ import annotations

import os
import shutil
import tempfile
import copy
import sqlite3
import sys
import subprocess
import traceback
import time
import threading
from datetime import datetime
from pathlib import Path
from functools import wraps
from pypdf import PdfReader, PdfWriter

from flask import Flask, flash, redirect, render_template, request, send_file, url_for, session, abort
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageChops
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from secrets import token_urlsafe, compare_digest
from datetime import timedelta

BASE = Path(__file__).resolve().parent
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
DB_PATH = os.environ.get("DSOUZA_DB_PATH", "").strip()
DB = Path(DB_PATH) if DB_PATH else BASE / "data" / "dsouza_orcamentos.db"
TEMPLATE_XLSX = BASE / "modelo" / "Orcamento_Modelo_DSouza.xlsx"
GENERATED = BASE / "data" / "gerados"
COMPANY_DIR = BASE / "data" / "empresa"
COMPANY_LOGO = COMPANY_DIR / "logo.png"
DEFAULT_LOGO = BASE / "modelo" / "logo_padrao.png"

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

IS_PRODUCTION = bool(os.environ.get("RENDER") or os.environ.get("DSOUZA_PRODUCTION"))

# Protecoes do cookie de sessao.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = IS_PRODUCTION
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)
app.config["SESSION_REFRESH_EACH_REQUEST"] = True

SECRET_KEY = os.environ.get("DSOUZA_SECRET_KEY", "").strip()

if IS_PRODUCTION and not SECRET_KEY:
    raise RuntimeError("DSOUZA_SECRET_KEY nao configurada no ambiente de producao.")

if not SECRET_KEY:
    SECRET_KEY = token_urlsafe(64)

if len(SECRET_KEY) < 32:
    raise RuntimeError("DSOUZA_SECRET_KEY deve ter pelo menos 32 caracteres.")

app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024

UNIDADES = ["m²", "cm²", "ml", "Kg", "g", "Un."]
PAGAMENTOS = ["PIX", "Crédito", "Débito", "Boleto", "Dinheiro"]
PRAZOS = ["Imediata", "5 dias", "7 dias", "15 dias", "21 dias", "30 dias", "45 dias", "60 dias"]
MAX_ITENS_TEMPLATE = 20
ALLOWED_LOGO = {"png", "jpg", "jpeg"}

DEFAULT_COMPANY = {
    "nome": "DSouza Interiores",
    "documento": "36.884.159/0001-63",
    "endereco": "Rua Hugo Van Der Goes, 348 - CEP: 02866-100",
    "email": "d.souzainteriores@gmail.com",
    "telefone": "11 95200-3076",
    "instagram": "dsouzainteriores",
}


class CompatRow(dict):
    """Linha com acesso por nome e tamb?m por ?ndice, como sqlite3.Row."""
    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.values())[key]
        return super().__getitem__(key)


def _pg_convert_row(cursor, row):
    if row is None:
        return None
    if not cursor.description:
        return row

    cols = []
    for desc in cursor.description:
        cols.append(getattr(desc, "name", desc[0]))

    return CompatRow(zip(cols, row))


def _pg_sql(query):
    query = query.replace(
        "printf('%06d', numero)",
        "LPAD(CAST(numero AS TEXT), 6, '0')"
    )
    return query.replace("?", "%s")


class PgCursorCompat:
    def __init__(self, cursor, lastrowid=None):
        self.cursor = cursor
        self.lastrowid = lastrowid

    def fetchone(self):
        return _pg_convert_row(self.cursor, self.cursor.fetchone())

    def fetchall(self):
        return [
            _pg_convert_row(self.cursor, row)
            for row in self.cursor.fetchall()
        ]

    def __iter__(self):
        for row in self.cursor:
            yield _pg_convert_row(self.cursor, row)


class PgConnectionCompat:
    def __init__(self, conn):
        self.conn = conn

    def execute(self, query, params=()):
        sql = _pg_sql(query)
        stripped = sql.strip().rstrip(";")
        upper = stripped.upper()

        # Em PostgreSQL, recupera automaticamente o ID inserido
        # para manter compatibilidade com cursor.lastrowid do SQLite.
        if upper.startswith("INSERT INTO") and " RETURNING " not in upper:
            sql = stripped + " RETURNING id"
            cur = self.conn.execute(sql, params)
            row = cur.fetchone()
            lastrowid = row[0] if row else None
            return PgCursorCompat(cur, lastrowid)

        cur = self.conn.execute(sql, params)
        return PgCursorCompat(cur)

    def executemany(self, sql, seq_of_params):
        sql = sql.replace("?", "%s")
        cur = self.conn.cursor()
        cur.executemany(sql, seq_of_params)
        return PgCursorCompat(cur)

    def executescript(self, script):
        # O schema SQLite usa AUTOINCREMENT.
        # SERIAL ? o equivalente apropriado para este projeto no PostgreSQL.
        script = script.replace(
            "INTEGER PRIMARY KEY AUTOINCREMENT",
            "SERIAL PRIMARY KEY"
        )

        for statement in script.split(";"):
            statement = statement.strip()
            if statement:
                self.conn.execute(statement)

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if exc_type is None:
                self.conn.commit()
            else:
                self.conn.rollback()
        finally:
            self.conn.close()


def db_conn():
    if DATABASE_URL:
        try:
            import psycopg
        except ImportError as exc:
            raise RuntimeError(
                "Driver PostgreSQL nao instalado. Instale psycopg[binary]."
            ) from exc

        conn = psycopg.connect(DATABASE_URL)
        return PgConnectionCompat(conn)

    DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


_LOGIN_ATTEMPTS = {}
_LOGIN_LOCK = threading.Lock()
_LOGIN_WINDOW = 15 * 60
_LOGIN_MAX_FAILURES = 5

def login_blocked(client_key):
    now = time.monotonic()
    with _LOGIN_LOCK:
        entry = _LOGIN_ATTEMPTS.get(client_key)
        if not entry:
            return False
        failures, first_at = entry
        if now - first_at >= _LOGIN_WINDOW:
            _LOGIN_ATTEMPTS.pop(client_key, None)
            return False
        return failures >= _LOGIN_MAX_FAILURES

def register_login_failure(client_key):
    now = time.monotonic()
    with _LOGIN_LOCK:
        failures, first_at = _LOGIN_ATTEMPTS.get(client_key, (0, now))
        if now - first_at >= _LOGIN_WINDOW:
            failures, first_at = 0, now
        _LOGIN_ATTEMPTS[client_key] = (failures + 1, first_at)

def clear_login_failures(client_key):
    with _LOGIN_LOCK:
        _LOGIN_ATTEMPTS.pop(client_key, None)

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped

def init_db():
    COMPANY_DIR.mkdir(parents=True, exist_ok=True)
    GENERATED.mkdir(parents=True, exist_ok=True)
    with db_conn() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user' CHECK(role IN ('admin','user')),
                ativo INTEGER NOT NULL DEFAULT 1,
                criado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS clientes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                documento TEXT,
                endereco TEXT,
                email TEXT,
                telefone TEXT,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS servicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                descricao TEXT NOT NULL UNIQUE,
                unidade TEXT,
                valor_padrao REAL,
                criado_em TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS orcamentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero INTEGER NOT NULL UNIQUE,
                cliente_id INTEGER,
                cliente_nome TEXT NOT NULL,
                cliente_documento TEXT,
                cliente_endereco TEXT,
                cliente_email TEXT,
                cliente_telefone TEXT,
                pagamento TEXT,
                prazo TEXT,
                observacoes TEXT,
                total REAL NOT NULL DEFAULT 0,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL,
                FOREIGN KEY (cliente_id) REFERENCES clientes(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS orcamento_itens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                orcamento_id INTEGER NOT NULL,
                descricao TEXT NOT NULL,
                quantidade REAL NOT NULL,
                unidade TEXT,
                valor_unitario REAL NOT NULL,
                total REAL NOT NULL,
                ordem INTEGER NOT NULL,
                FOREIGN KEY (orcamento_id) REFERENCES orcamentos(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS empresa_config (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                nome TEXT,
                documento TEXT,
                endereco TEXT,
                email TEXT,
                telefone TEXT,
                instagram TEXT,
                logo_path TEXT,
                atualizado_em TEXT NOT NULL
            );
            """
        )
        row = con.execute("SELECT id FROM empresa_config WHERE id=1").fetchone()
        if not row:
            con.execute(
                """INSERT INTO empresa_config
                (id,nome,documento,endereco,email,telefone,instagram,logo_path,atualizado_em)
                VALUES(1,?,?,?,?,?,?,?,?)""",
                (
                    DEFAULT_COMPANY["nome"], DEFAULT_COMPANY["documento"], DEFAULT_COMPANY["endereco"],
                    DEFAULT_COMPANY["email"], DEFAULT_COMPANY["telefone"], DEFAULT_COMPANY["instagram"],
                    "", datetime.now().isoformat(timespec="seconds"),
                ),
            )


def get_company():
    with db_conn() as con:
        row = con.execute("SELECT * FROM empresa_config WHERE id=1").fetchone()
    return row


def next_number(con):
    row = con.execute("SELECT COALESCE(MAX(numero), 0) + 1 AS n FROM orcamentos").fetchone()
    return int(row["n"])


def as_float(value, default=0.0):
    try:
        if isinstance(value, str):
            value = value.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        return float(value)
    except (TypeError, ValueError):
        return default


def save_budget_from_form():
    cliente_id = request.form.get("cliente_id") or None
    nome = (request.form.get("cliente_nome") or "").strip()
    if not nome:
        raise ValueError("Informe o nome do cliente.")

    documento = (request.form.get("cliente_documento") or "").strip()
    endereco = (request.form.get("cliente_endereco") or "").strip()
    email = (request.form.get("cliente_email") or "").strip()
    telefone = (request.form.get("cliente_telefone") or "").strip()
    pagamento = (request.form.get("pagamento") or "").strip()
    prazo = (request.form.get("prazo") or "").strip()
    observacoes = (request.form.get("observacoes") or "").strip()

    descricoes = request.form.getlist("item_descricao[]")
    quantidades = request.form.getlist("item_quantidade[]")
    unidades = request.form.getlist("item_unidade[]")
    valores = request.form.getlist("item_valor[]")

    itens = []
    for i, desc in enumerate(descricoes):
        desc = (desc or "").strip()
        qtd = as_float(quantidades[i] if i < len(quantidades) else 0)
        valor = as_float(valores[i] if i < len(valores) else 0)
        unidade = (unidades[i] if i < len(unidades) else "").strip()
        if not desc and qtd == 0 and valor == 0:
            continue
        if not desc:
            raise ValueError(f"Informe a descrição do item {i + 1}.")
        if qtd <= 0:
            raise ValueError(f"A quantidade do item {i + 1} deve ser maior que zero.")
        if valor < 0:
            raise ValueError(f"O valor do item {i + 1} não pode ser negativo.")
        itens.append({"descricao": desc, "quantidade": qtd, "unidade": unidade, "valor": valor, "total": qtd * valor})

    if not itens:
        raise ValueError("Adicione pelo menos um item ao orçamento.")
    if len(itens) > MAX_ITENS_TEMPLATE:
        raise ValueError(f"O modelo original comporta até {MAX_ITENS_TEMPLATE} itens por orçamento.")

    total = sum(i["total"] for i in itens)
    now = datetime.now().isoformat(timespec="seconds")

    with db_conn() as con:
        if cliente_id:
            existing = con.execute("SELECT id FROM clientes WHERE id=?", (cliente_id,)).fetchone()
            if existing:
                con.execute(
                    "UPDATE clientes SET nome=?, documento=?, endereco=?, email=?, telefone=? WHERE id=?",
                    (nome, documento, endereco, email, telefone, cliente_id),
                )
            else:
                cliente_id = None
        if not cliente_id:
            cur = con.execute(
                "INSERT INTO clientes(nome, documento, endereco, email, telefone) VALUES(?,?,?,?,?)",
                (nome, documento, endereco, email, telefone),
            )
            cliente_id = cur.lastrowid

        numero = next_number(con)
        cur = con.execute(
            """INSERT INTO orcamentos(
                numero, cliente_id, cliente_nome, cliente_documento, cliente_endereco,
                cliente_email, cliente_telefone, pagamento, prazo, observacoes,
                total, criado_em, atualizado_em
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (numero, cliente_id, nome, documento, endereco, email, telefone, pagamento, prazo,
             observacoes, total, now, now),
        )
        oid = cur.lastrowid
        con.executemany(
            "INSERT INTO orcamento_itens(orcamento_id, descricao, quantidade, unidade, valor_unitario, total, ordem) VALUES(?,?,?,?,?,?,?)",
            [(oid, it["descricao"], it["quantidade"], it["unidade"], it["valor"], it["total"], idx)
             for idx, it in enumerate(itens, 1)],
        )
    return oid


def update_budget_from_form(orcamento_id):
    cliente_id = request.form.get("cliente_id") or None
    nome = (request.form.get("cliente_nome") or "").strip()
    if not nome:
        raise ValueError("Informe o nome do cliente.")

    documento = (request.form.get("cliente_documento") or "").strip()
    endereco = (request.form.get("cliente_endereco") or "").strip()
    email = (request.form.get("cliente_email") or "").strip()
    telefone = (request.form.get("cliente_telefone") or "").strip()
    pagamento = (request.form.get("pagamento") or "").strip()
    prazo = (request.form.get("prazo") or "").strip()
    observacoes = (request.form.get("observacoes") or "").strip()

    descricoes = request.form.getlist("item_descricao[]")
    quantidades = request.form.getlist("item_quantidade[]")
    unidades = request.form.getlist("item_unidade[]")
    valores = request.form.getlist("item_valor[]")

    itens = []
    for i, desc in enumerate(descricoes):
        desc = (desc or "").strip()
        qtd = as_float(quantidades[i] if i < len(quantidades) else 0)
        valor = as_float(valores[i] if i < len(valores) else 0)
        unidade = (unidades[i] if i < len(unidades) else "").strip()
        if not desc and qtd == 0 and valor == 0:
            continue
        if not desc:
            raise ValueError(f"Informe a descrição do item {i + 1}.")
        if qtd <= 0:
            raise ValueError(f"A quantidade do item {i + 1} deve ser maior que zero.")
        if valor < 0:
            raise ValueError(f"O valor do item {i + 1} não pode ser negativo.")
        itens.append({"descricao": desc, "quantidade": qtd, "unidade": unidade, "valor": valor, "total": qtd * valor})

    if not itens:
        raise ValueError("Adicione pelo menos um item ao orçamento.")
    if len(itens) > MAX_ITENS_TEMPLATE:
        raise ValueError(f"O modelo original comporta até {MAX_ITENS_TEMPLATE} itens por orçamento.")

    total = sum(i["total"] for i in itens)
    now = datetime.now().isoformat(timespec="seconds")

    with db_conn() as con:
        atual = con.execute("SELECT * FROM orcamentos WHERE id=?", (orcamento_id,)).fetchone()
        if not atual:
            raise FileNotFoundError("Orçamento não encontrado.")

        if cliente_id:
            existing = con.execute("SELECT id FROM clientes WHERE id=?", (cliente_id,)).fetchone()
            if existing:
                con.execute(
                    "UPDATE clientes SET nome=?, documento=?, endereco=?, email=?, telefone=? WHERE id=?",
                    (nome, documento, endereco, email, telefone, cliente_id),
                )
            else:
                cliente_id = None
        if not cliente_id:
            cur = con.execute(
                "INSERT INTO clientes(nome, documento, endereco, email, telefone) VALUES(?,?,?,?,?)",
                (nome, documento, endereco, email, telefone),
            )
            cliente_id = cur.lastrowid

        con.execute(
            """UPDATE orcamentos SET
               cliente_id=?, cliente_nome=?, cliente_documento=?, cliente_endereco=?,
               cliente_email=?, cliente_telefone=?, pagamento=?, prazo=?, observacoes=?,
               total=?, atualizado_em=? WHERE id=?""",
            (cliente_id, nome, documento, endereco, email, telefone, pagamento, prazo,
             observacoes, total, now, orcamento_id),
        )
        con.execute("DELETE FROM orcamento_itens WHERE orcamento_id=?", (orcamento_id,))
        con.executemany(
            "INSERT INTO orcamento_itens(orcamento_id, descricao, quantidade, unidade, valor_unitario, total, ordem) VALUES(?,?,?,?,?,?,?)",
            [(orcamento_id, it["descricao"], it["quantidade"], it["unidade"], it["valor"], it["total"], idx)
             for idx, it in enumerate(itens, 1)],
        )
    return orcamento_id


def budget_data(orcamento_id):
    with db_conn() as con:
        o = con.execute("SELECT * FROM orcamentos WHERE id=?", (orcamento_id,)).fetchone()
        if not o:
            return None, []
        itens = con.execute("SELECT * FROM orcamento_itens WHERE orcamento_id=? ORDER BY ordem", (orcamento_id,)).fetchall()
        return o, itens


def _prepare_logo_for_excel(source: Path) -> Path:
    """Recorta margens transparentes/brancas para não achatar a marca no PDF."""
    GENERATED.mkdir(parents=True, exist_ok=True)
    out = GENERATED / "_logo_render.png"
    with PILImage.open(source) as original:
        image = original.convert("RGBA")
        full = (0, 0, image.width, image.height)
        bbox = image.getchannel("A").getbbox()

        # JPGs e imagens sem transparência: detecta conteúdo diferente do branco.
        if not bbox or bbox == full:
            rgb = image.convert("RGB")
            bg = PILImage.new("RGB", rgb.size, "white")
            diff = ImageChops.difference(rgb, bg).convert("L")
            diff = diff.point(lambda p: 255 if p > 18 else 0)
            bbox = diff.getbbox()

        if bbox:
            left, top, right, bottom = bbox
            pad = max(4, int(min(image.width, image.height) * 0.015))
            left = max(0, left - pad)
            top = max(0, top - pad)
            right = min(image.width, right + pad)
            bottom = min(image.height, bottom + pad)
            image = image.crop((left, top, right, bottom))

        image.save(out, "PNG", optimize=True)
    return out


def _arrange_header_images(ws, logo_source: Path):
    """Alinha logo e ícones do cabeçalho próximos aos respectivos textos."""
    # Identifica os desenhos originais pela âncora do modelo.
    icons = {}
    kept = []
    for img in list(ws._images):
        try:
            anchor = img.anchor._from
            key = (anchor.row, anchor.col)
        except Exception:
            kept.append(img)
            continue

        if key == (0, 0):       # logo antigo
            continue
        if key == (0, 3):       # CNPJ
            icons["documento"] = img
        elif key == (1, 1):     # localização
            icons["endereco"] = img
        elif key == (2, 3):     # e-mail
            icons["email"] = img
        elif key == (3, 4):     # WhatsApp
            icons["telefone"] = img
        else:
            kept.append(img)

    ws._images = kept

    # Logo sempre é reinserida recortada, inclusive a logo padrão do modelo.
    render_logo = _prepare_logo_for_excel(logo_source)
    with PILImage.open(render_logo) as pil:
        w, h = pil.size
    max_w, max_h = 350, 120
    scale = min(max_w / max(w, 1), max_h / max(h, 1))
    logo = XLImage(str(render_logo))
    logo.width = max(1, int(w * scale))
    logo.height = max(1, int(h * scale))
    ws.add_image(logo, "A1")

    # Ãcones ficam no início do próprio bloco de texto para evitar o grande vão visual.
    # O texto recebe um pequeno recuo para não ficar sobre a imagem.
    positions = {
    	"documento": ("D1", 17, 17),
    	"endereco": ("D2", 18, 18),
    	"email": ("D3", 18, 18),
    	   "telefone": ("D4", 18, 18),
}
    for name, (cell, width, height) in positions.items():
        img = icons.get(name)
        if img is not None:
            img.anchor = cell
            img.width = width
            img.height = height
            ws.add_image(img)

def validate_template_structure():
    """Valida os pontos essenciais do Excel mestre antes de gerar uma cópia."""
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError("A planilha modelo não foi encontrada.")

    check = load_workbook(TEMPLATE_XLSX, read_only=False, data_only=False)
    try:
        if "ORÇAMENTO" not in check.sheetnames:
            raise RuntimeError("O modelo mestre está inválido: aba ORÇAMENTO não encontrada.")
        ws = check["ORÇAMENTO"]
        expected = {
            "A8": "DADOS DO CLIENTE",
            "A14": "ORÇAMENTO",
            "A15": "Item",
            "B15": "Qtd.",
            "C15": "Uni.",
            "D15": "Valor Item",
            "E15": "Total Item",
        }
        for cell, value in expected.items():
            if ws[cell].value != value:
                raise RuntimeError(
                    f"O modelo mestre parece ter sido alterado ({cell}). "
                    "Restaure o arquivo do sistema antes de gerar novos orçamentos."
                )

        prefixes = {
            "C9": "CPF/CNPJ:",
            "C11": "Telefone:",
            "D38": "Condições de pagamento:",
            "D39": "Prazo de entrega:",
        }
        for cell, prefix in prefixes.items():
            current = str(ws[cell].value or "").strip()
            if not current.startswith(prefix):
                raise RuntimeError(
                    f"O modelo mestre parece ter sido alterado ({cell}). "
                    "Restaure o arquivo do sistema antes de gerar novos orçamentos."
                )

        if ws["A16"].font.name != "Arial":
            raise RuntimeError(
                "A fonte principal do modelo mestre foi alterada. "
                "Restaure o modelo do sistema antes de gerar novos orçamentos."
            )
    finally:
        check.close()

def generate_excel(orcamento_id):
    o, itens = budget_data(orcamento_id)
    if not o:
        raise FileNotFoundError("Orçamento não encontrado.")
    validate_template_structure()

    GENERATED.mkdir(parents=True, exist_ok=True)
    out = GENERATED / f"Orcamento_{int(o['numero']):06d}_{safe_name(o['cliente_nome'])}.xlsx"
    shutil.copy2(TEMPLATE_XLSX, out)

    wb = load_workbook(out)
    ws = wb["ORÇAMENTO"]

    # Ajuste da coluna ITEM para evitar corte no PDF.
    # Proporções compactas para o PDF: prioriza a descrição e evita grandes
    # vazios nas colunas curtas de quantidade, unidade e valores.
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    empresa = get_company()

    # Cabeçalho compacto: ícone e texto ficam no mesmo bloco visual.
    if empresa:
        recuo = "      "
        ws["D1"] = f"{recuo}{empresa['documento'] or ''}"
        ws["D2"] = f"{recuo}{empresa['endereco'] or ''}"
        ws["D3"] = f"{recuo}{empresa['email'] or ''}"
        ws["D4"] = f"{recuo}{empresa['telefone'] or ''}"

        custom = BASE / empresa["logo_path"] if empresa["logo_path"] else None
        logo_source = custom if custom and custom.exists() else DEFAULT_LOGO
    else:
        logo_source = DEFAULT_LOGO

    if logo_source.exists():
        _arrange_header_images(ws, logo_source)

    # Cliente: rótulo e valor juntos, usando a largura de C:D para não criar vão nem corte.
    for faixa in ("C9:D9", "C11:D11"):
        if faixa not in [str(rng) for rng in ws.merged_cells.ranges]:
            ws.merge_cells(faixa)
    ws["A9"] = f"Nome:  {o['cliente_nome'] or ''}"
    ws["C9"] = f"CPF/CNPJ:  {o['cliente_documento'] or ''}"
    ws["A10"] = f"Endereço:  {o['cliente_endereco'] or ''}"
    ws["A11"] = f"E-mail:  {o['cliente_email'] or ''}"
    ws["C11"] = f"Telefone:  {o['cliente_telefone'] or ''}"

    # Limpa as linhas de itens. Linhas sem item ficam realmente vazias.
    # O formato do total também oculta zero, evitando R$ 0,00 em linhas não utilizadas.
    for row in range(16, 36):
        ws.cell(row, 1).value = None
        ws.cell(row, 2).value = None
        ws.cell(row, 3).value = None
        ws.cell(row, 4).value = None
        ws.cell(row, 5).value = None
        ws.cell(row, 5).number_format = '[$R$-pt-BR] #.##0,00;[Red]-[$R$-pt-BR] #.##0,00;;'

    for idx, it in enumerate(itens, start=16):
        ws.cell(idx, 1).value = it["descricao"]
        ws.cell(idx, 2).value = it["quantidade"]
        ws.cell(idx, 3).value = it["unidade"]
        ws.cell(idx, 4).value = it["valor_unitario"]
        ws.cell(idx, 5).value = it["total"]

    used_rows = set(range(16, 16 + len(itens)))
    for row in range(16, 36):
        if row in used_rows:
            ws.row_dimensions[row].hidden = False
            description = str(ws.cell(row, 1).value or "")
            if description:
                # Conta separadamente quebras manuais e quebras automáticas.
                # Somar apenas o tamanho total cortava a última linha no Excel.
                linhas = sum(
                    max(1, (len(parte) + 54) // 55)
                    for parte in (description.splitlines() or [""])
                )
                altura = min(180, max(30, linhas * 15))
                ws.row_dimensions[row].height = altura
            else:
                ws.row_dimensions[row].height = 22
        else:
            # As linhas existem para permitir até 20 itens, mas não devem
            # aparecer vazias no Excel/PDF quando o orçamento usa menos.
            ws.row_dimensions[row].height = 0
            ws.row_dimensions[row].hidden = True

    ws["E36"] = o["total"]

    # Condições e prazo ficam em um único bloco, sem o grande espaço entre rótulo e valor.
    for faixa in ("D38:E38", "D39:E39"):
        if faixa not in [str(rng) for rng in ws.merged_cells.ranges]:
            ws.merge_cells(faixa)
    ws["D38"] = f"Condições de pagamento:  {o['pagamento'] or ''}"
    ws["D39"] = f"Prazo de entrega:  {o['prazo'] or ''}"
    # Observações usam uma área larga e alta para não cortar textos com mais de uma linha.
    faixa_obs = "A41:E44"
    if faixa_obs not in [str(rng) for rng in ws.merged_cells.ranges]:
        # Remove mesclagens conflitantes na área, se houver.
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= 44 and rng.max_row >= 41 and rng.min_col <= 5 and rng.max_col >= 1:
                ws.unmerge_cells(str(rng))
        ws.merge_cells(faixa_obs)
    obs_texto = (o["observacoes"] or "").strip()
    ws["A41"] = f"OBS: {obs_texto}" if obs_texto else "OBS:"
    ws["A41"].alignment = copy.copy(ws["A41"].alignment)
    ws["A41"].alignment = ws["A41"].alignment.copy(wrap_text=True, vertical="top", horizontal="left")
    linhas_texto = obs_texto.splitlines() or [""]
    linhas_visuais = sum(max(1, (len(linha) + 89) // 90) for linha in linhas_texto)
    altura_total = 72 if not obs_texto else min(120, max(72, 18 + linhas_visuais * 13))
    altura_linha = altura_total / 4
    for row in range(41, 45):
        ws.row_dimensions[row].height = altura_linha

      # Mensagens finais: validade em cima e agradecimento logo abaixo.
    for faixa in ("A45:E45", "A47:E47"):
        for rng in list(ws.merged_cells.ranges):
            if str(rng) == faixa:
                break
        else:
            ws.merge_cells(faixa)

    ws["A45"] = 'O orçamento é válido por 30 dias.'
    ws["A45"].font = copy.copy(ws["A41"].font)
    ws["A45"].font = ws["A45"].font.copy(
        name="Arial", size=12, bold=True, italic=False
    )
    ws["A45"].alignment = copy.copy(ws["A41"].alignment)
    ws["A45"].alignment = ws["A45"].alignment.copy(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[45].height = 24

    closing = 'Desde já, agradecemos a preferência.'
    ws["A47"] = closing
    ws["A47"].font = copy.copy(ws["A41"].font)
    ws["A47"].font = ws["A47"].font.copy(
        name="Arial", size=11, bold=False, italic=True
    )
    ws["A47"].alignment = copy.copy(ws["A41"].alignment)
    ws["A47"].alignment = ws["A47"].alignment.copy(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[47].height = 24

    # Limpa a frase antiga do modelo para não duplicar.
    if ws["C44"].value == closing:
        ws["C44"] = None

    ws.print_area = "A1:E47"
    ws.print_title_rows = "14:15"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.20
    ws.page_margins.right = 0.20
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    for row in range(16, 36):
        ws.cell(row, 1).alignment = copy.copy(ws.cell(row, 1).alignment)
        ws.cell(row, 1).alignment = ws.cell(row, 1).alignment.copy(
            horizontal="left",
            vertical="center",
            wrap_text=True,
            indent=1,
            shrink_to_fit=False,
        )
        for col in (4, 5):
            ws.cell(row, col).alignment = copy.copy(ws.cell(row, col).alignment)
            ws.cell(row, col).alignment = ws.cell(row, col).alignment.copy(
                horizontal="center", vertical="center"
            )
        ws.cell(row, 5).font = copy.copy(ws.cell(row, 5).font)
        ws.cell(row, 5).font = ws.cell(row, 5).font.copy(bold=True)

    wb.save(out)
    return out

def safe_name(text):
    allowed = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_"
    clean = "_".join((text or "cliente").split())
    clean = "".join(c for c in clean if c in allowed)
    return clean[:40] or "cliente"




def libreoffice_to_pdf(xlsx_path: Path):
    """Converte XLSX para PDF usando LibreOffice em modo headless."""
    xlsx_path = xlsx_path.resolve()

    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Arquivo Excel nao encontrado: {xlsx_path.name}"
        )

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError(
            "LibreOffice nao encontrado no servidor."
        )

    pdf_path = xlsx_path.with_suffix(".pdf")

    # Evita interpretar um PDF antigo como uma conversao nova.
    if pdf_path.exists():
        pdf_path.unlink()

    resultado = subprocess.run(
        [
            soffice,
            f"-env:UserInstallation={Path(tempfile.mkdtemp()).resolve().as_uri()}",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(xlsx_path.parent),
            str(xlsx_path),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=120,
    )

    if resultado.returncode != 0:
        detalhe = (resultado.stderr or resultado.stdout or "").strip()
        raise RuntimeError(
            f"LibreOffice falhou ao gerar o PDF. Detalhe: {detalhe}"
        )

    if not pdf_path.exists() or pdf_path.stat().st_size == 0:
        detalhe = (resultado.stdout or resultado.stderr or "").strip()
        raise RuntimeError(
            f"LibreOffice terminou sem criar o PDF. Detalhe: {detalhe}"
        )

    # Remove paginas auxiliares que o LibreOffice pode exportar.
    reader = PdfReader(str(pdf_path))
    if len(reader.pages) > 1:
        writer = PdfWriter()
        writer.add_page(reader.pages[0])
        temp_pdf = pdf_path.with_name(pdf_path.stem + "_pag1.pdf")
        with temp_pdf.open("wb") as f:
            writer.write(f)
        temp_pdf.replace(pdf_path)

    return pdf_path

def excel_to_pdf(xlsx_path: Path):
    """Converte o Excel gerado para PDF usando o Microsoft Excel no Windows.

    Cada requisição do Flask pode rodar em uma thread diferente. O COM do Excel
    precisa ser inicializado explicitamente nessa thread para funcionar de forma
    estável. A configuração de página já vem pronta do arquivo .xlsx, então não
    alteramos PageSetup pelo COM antes da exportação.
    """
    if sys.platform != "win32":
        return libreoffice_to_pdf(xlsx_path)

    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "O componente de integração com o Excel não está instalado. "
            "Execute INSTALAR_E_INICIAR.bat novamente."
        ) from exc

    xlsx_path = xlsx_path.resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {xlsx_path.name}")

    pdf_path = xlsx_path.with_suffix(".pdf")
    if pdf_path.exists():
        try:
            pdf_path.unlink()
        except OSError:
            pass

    excel = None
    workbook = None
    com_initialized = False
    try:
        # Flask atende requisições em threads; o COM precisa ser iniciado em cada uma.
        pythoncom.CoInitialize()
        com_initialized = True

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        workbook = excel.Workbooks.Open(
            Filename=str(xlsx_path),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        worksheet = workbook.Worksheets("ORÇAMENTO")

        # O print area, A4, margens e ajuste para uma página já estão salvos no modelo.
        # Alterar PageSetup via COM pode falhar em algumas instalações do Excel/impressora.
        worksheet.ExportAsFixedFormat(
            Type=0,
            Filename=str(pdf_path),
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )

        if not pdf_path.exists() or pdf_path.stat().st_size == 0:
            raise RuntimeError("O Excel terminou a exportação, mas o arquivo PDF não foi criado.")

        return pdf_path
    except Exception as exc:
        # Registra o erro real no terminal para diagnóstico, sem mascará-lo.
        print("\n[ERRO PDF] Falha ao converter Excel para PDF:", file=sys.stderr)
        traceback.print_exc()
        detail = str(exc).strip() or exc.__class__.__name__
        raise RuntimeError(f"Não foi possível gerar o PDF pelo Excel. Detalhe: {detail}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if com_initialized:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass




def csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = token_urlsafe(32)
    return session["csrf_token"]

app.jinja_env.globals["csrf_token"] = csrf_token


@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"

    if IS_PRODUCTION:
        response.headers["Strict-Transport-Security"] = (
            "max-age=31536000; includeSubDomains"
        )

    return response


@app.before_request
def security_before_request():
    if request.method == "POST":
        expected = session.get("csrf_token")
        supplied = request.form.get("csrf_token", "")

        if not expected or not supplied or not compare_digest(expected, supplied):
            abort(
                400,
                description="Token de seguranca invalido. Atualize a pagina e tente novamente."
            )


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("novo_orcamento"))

    if "csrf_token" not in session:
        session["csrf_token"] = token_urlsafe(32)

    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")

        if len(username) > 40 or len(password) > 256:
            flash("Usuário ou senha inválidos.", "error")
            return render_template("login.html")

        client_key = request.remote_addr or "unknown"

        if login_blocked(client_key):
            flash("Muitas tentativas. Aguarde 15 minutos e tente novamente.", "error")
            return render_template("login.html"), 429

        with db_conn() as conn:
            user = conn.execute(
                "SELECT * FROM usuarios WHERE username=? AND ativo=1",
                (username,)
            ).fetchone()

        if not user or not check_password_hash(user["password_hash"], password):
            register_login_failure(client_key)
            flash("Usuário ou senha inválidos.", "error")
            return render_template("login.html")

        clear_login_failures(client_key)
        session.clear()
        session.permanent = True
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user["role"]
        session["csrf_token"] = token_urlsafe(32)

        next_url = request.args.get("next") or request.form.get("next") or url_for("novo_orcamento")
        if not next_url.startswith("/") or next_url.startswith("//"):
            next_url = url_for("novo_orcamento")

        return redirect(next_url)

    return render_template("login.html")


@app.post("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/", methods=["GET", "POST"])
@login_required
def novo_orcamento():
    if request.method == "POST":
        try:
            oid = save_budget_from_form()
            flash("Orçamento salvo com sucesso.", "success")
            return redirect(url_for("ver_orcamento", orcamento_id=oid))
        except ValueError as e:
            flash(str(e), "error")

    with db_conn() as con:
        clientes = con.execute("SELECT * FROM clientes ORDER BY nome").fetchall()
        servicos = con.execute("SELECT * FROM servicos ORDER BY descricao").fetchall()
    return render_template(
        "novo.html", clientes=clientes, servicos=servicos,
        unidades=UNIDADES, pagamentos=PAGAMENTOS, prazos=PRAZOS,
        max_itens=MAX_ITENS_TEMPLATE,
    )


@app.route("/clientes", methods=["GET", "POST"])
@login_required
def clientes():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        if not nome:
            flash("Informe o nome do cliente.", "error")
        else:
            with db_conn() as con:
                con.execute(
                    "INSERT INTO clientes(nome, documento, endereco, email, telefone) VALUES(?,?,?,?,?)",
                    (nome, request.form.get("documento", "").strip(), request.form.get("endereco", "").strip(),
                     request.form.get("email", "").strip(), request.form.get("telefone", "").strip()),
                )
            flash("Cliente cadastrado.", "success")
            return redirect(url_for("clientes"))
    with db_conn() as con:
        rows = con.execute("SELECT * FROM clientes ORDER BY nome").fetchall()
    return render_template("clientes.html", clientes=rows)


@app.post("/clientes/<int:cliente_id>/excluir")
@login_required
def excluir_cliente(cliente_id):
    with db_conn() as con:
        con.execute("DELETE FROM clientes WHERE id=?", (cliente_id,))
    flash("Cliente removido.", "success")
    return redirect(url_for("clientes"))


@app.route("/servicos", methods=["GET", "POST"])
@login_required
def servicos():
    if request.method == "POST":
        desc = (request.form.get("descricao") or "").strip()
        if not desc:
            flash("Informe a descrição do serviço.", "error")
        else:
            try:
                with db_conn() as con:
                    con.execute(
                        "INSERT INTO servicos(descricao, unidade, valor_padrao) VALUES(?,?,?)",
                        (desc, request.form.get("unidade", "").strip(), as_float(request.form.get("valor_padrao"))),
                    )
                flash("Serviço cadastrado.", "success")
                return redirect(url_for("servicos"))
            except sqlite3.IntegrityError:
                flash("Esse serviço já está cadastrado.", "error")
    with db_conn() as con:
        rows = con.execute("SELECT * FROM servicos ORDER BY descricao").fetchall()
    return render_template("servicos.html", servicos=rows, unidades=UNIDADES)


@app.post("/servicos/<int:servico_id>/excluir")
@login_required
def excluir_servico(servico_id):
    with db_conn() as con:
        con.execute("DELETE FROM servicos WHERE id=?", (servico_id,))
    flash("Serviço removido.", "success")
    return redirect(url_for("servicos"))


@app.route("/empresa", methods=["GET", "POST"])
@login_required
def empresa():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        documento = (request.form.get("documento") or "").strip()
        endereco = (request.form.get("endereco") or "").strip()
        email = (request.form.get("email") or "").strip()
        telefone = (request.form.get("telefone") or "").strip()
        instagram = (request.form.get("instagram") or "").strip().lstrip("@")

        logo_rel = None
        current = get_company()
        if current:
            logo_rel = current["logo_path"] or ""

        logo = request.files.get("logo")
        if logo and logo.filename:
            filename = secure_filename(logo.filename)
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if ext not in ALLOWED_LOGO:
                flash("Logo inválido. Use PNG, JPG ou JPEG.", "error")
                return redirect(url_for("empresa"))
            try:
                COMPANY_DIR.mkdir(parents=True, exist_ok=True)
                with PILImage.open(logo.stream) as image:
                    image = image.convert("RGBA")
                    image.thumbnail((1600, 1600))
                    image.save(COMPANY_LOGO, "PNG", optimize=True)
                logo_rel = str(COMPANY_LOGO.relative_to(BASE)).replace("\\", "/")
            except Exception:
                flash("Não foi possível processar essa imagem. Tente outro arquivo.", "error")
                return redirect(url_for("empresa"))

        with db_conn() as con:
            con.execute(
                """UPDATE empresa_config SET nome=?,documento=?,endereco=?,email=?,telefone=?,instagram=?,logo_path=?,atualizado_em=? WHERE id=1""",
                (nome, documento, endereco, email, telefone, instagram, logo_rel,
                 datetime.now().isoformat(timespec="seconds")),
            )
        flash("Informações da empresa atualizadas.", "success")
        return redirect(url_for("empresa"))

    return render_template("empresa.html", empresa=get_company(), has_logo=COMPANY_LOGO.exists())


@app.post("/empresa/logo/remover")
@login_required
def remover_logo():
    if COMPANY_LOGO.exists():
        COMPANY_LOGO.unlink()
    with db_conn() as con:
        con.execute("UPDATE empresa_config SET logo_path='', atualizado_em=? WHERE id=1",
                    (datetime.now().isoformat(timespec="seconds"),))
    flash("Logo personalizada removida. Novos orçamentos voltarão a usar a marca original da planilha.", "success")
    return redirect(url_for("empresa"))


@app.get("/empresa/logo")
@login_required
def ver_logo_empresa():
    if not COMPANY_LOGO.exists():
        return "Logo não cadastrada", 404
    return send_file(COMPANY_LOGO)


@app.get("/historico")
@login_required
def historico():
    q = (request.args.get("q") or "").strip()
    with db_conn() as con:
        if q:
            rows = con.execute(
                "SELECT * FROM orcamentos WHERE cliente_nome LIKE ? OR printf('%06d', numero) LIKE ? ORDER BY numero DESC",
                (f"%{q}%", f"%{q}%"),
            ).fetchall()
        else:
            rows = con.execute("SELECT * FROM orcamentos ORDER BY numero DESC").fetchall()
    return render_template("historico.html", orcamentos=rows, q=q)


@app.get("/orcamentos/<int:orcamento_id>")
@login_required
def ver_orcamento(orcamento_id):
    o, itens = budget_data(orcamento_id)
    if not o:
        return "Orçamento não encontrado", 404
    return render_template("detalhe.html", o=o, itens=itens)


@app.route("/orcamentos/<int:orcamento_id>/editar", methods=["GET", "POST"])
@login_required
def editar_orcamento(orcamento_id):
    o, itens = budget_data(orcamento_id)
    if not o:
        return "Orçamento não encontrado", 404

    if request.method == "POST":
        try:
            update_budget_from_form(orcamento_id)
            flash("Orçamento atualizado com sucesso.", "success")
            return redirect(url_for("ver_orcamento", orcamento_id=orcamento_id))
        except (ValueError, FileNotFoundError) as e:
            flash(str(e), "error")
            o, itens = budget_data(orcamento_id)

    with db_conn() as con:
        clientes = con.execute("SELECT * FROM clientes ORDER BY nome").fetchall()
        servicos = con.execute("SELECT * FROM servicos ORDER BY descricao").fetchall()

    return render_template(
        "editar.html", o=o, itens=itens, clientes=clientes, servicos=servicos,
        unidades=UNIDADES, pagamentos=PAGAMENTOS, prazos=PRAZOS,
        max_itens=MAX_ITENS_TEMPLATE,
    )


@app.get("/orcamentos/<int:orcamento_id>/excel")
@login_required
def baixar_excel(orcamento_id):
    try:
        path = generate_excel(orcamento_id)
        return send_file(path, as_attachment=True, download_name=path.name)
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("ver_orcamento", orcamento_id=orcamento_id))


@app.get("/orcamentos/<int:orcamento_id>/pdf")
@login_required
def baixar_pdf(orcamento_id):
    try:
        xlsx = generate_excel(orcamento_id)
        pdf = excel_to_pdf(xlsx)
        return send_file(pdf, as_attachment=True, download_name=pdf.name)
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("ver_orcamento", orcamento_id=orcamento_id))


@app.post("/orcamentos/<int:orcamento_id>/excluir")
@login_required
def excluir_orcamento(orcamento_id):
    with db_conn() as con:
        con.execute("DELETE FROM orcamentos WHERE id=?", (orcamento_id,))
    flash("Orçamento excluído do histórico.", "success")
    return redirect(url_for("historico"))


@app.template_filter("brl")
def brl(value):
    value = float(value or 0)
    s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


@app.template_filter("data_br")
def data_br(value):
    try:
        return datetime.fromisoformat(value).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return value


if __name__ == "__main__":
    init_db()
    host = os.environ.get("DSOUZA_HOST", "127.0.0.1")
    port = int(os.environ.get("DSOUZA_PORT", "5000"))
    print(f"\nDSouza Orçamentos: http://{host}:{port}\n")
    app.run(host=host, port=port, debug=False)

















